import face_recognition
import cv2
import os
import glob
import numpy as np
from simple_facerec import SimpleFacerec
import RPi.GPIO as GPIO
from datetime import datetime
from datetime import date
import pandas as pd
import openpyxl

def lock(pin):
    GPIO.output(pin, GPIO.LOW)
    print('Kişi Tanımlı Değil')


# Kilit Aç
def unlock(pin):
    GPIO.output(pin, GPIO.HIGH)
    print('Kişi Tanımlı')

#GPIO26 pinini seçtiğimizi belirtiyoruz
role_pini = [17]
pin=[11]
# Gelebilecek gereksiz uyarıları devre dışı bırakıyoruz
GPIO.setwarnings(False)
# GPIO numaralarına göre seçim yaptık
GPIO.setmode(GPIO.BCM)
GPIO.setup(pin, GPIO.IN)    
# Seçtiğimiz röle pinini çıkış olarak ayarlıyoruz.
GPIO.setup(role_pini, GPIO.OUT)

    
today = date.today()
day = today.strftime("%b-%d-%Y")
day_str = "yoklama-" + day + ".csv"
print(day_str)

dosya = open(day_str, "a")
dosya.write("Ad, Saat")
dosya.close()
def yoklamayaYaz(name):
    #with open('yoklama.csv','r+') as f:
    with open(day_str, 'r+') as f:
        myDataList = f.readlines()
        nameList = []
        for line in myDataList:
            entry = line.split(',')
            nameList.append(entry[0])

        if name not in nameList:
            now = datetime.now()
            dtString = now.strftime('%H:%M:%S')
            f.writelines(f'\n{name},{dtString}')


sfr=SimpleFacerec()
sfr.load_encoding_images("images/")
cap=cv2.VideoCapture(1)
while True:
	ret,frame=cap.read()
	face_locations,face_names=sfr.detect_known_faces(frame)
	for face_loc, name in zip(face_locations,face_names):
		y1, x1, y2, x2=face_loc[0], face_loc[1],face_loc[2],face_loc[3]
		if name=="Unknown":
			lock(role_pini)
		else:
			unlock(role_pini)
			yoklamayaYaz(name)
		cv2.putText(frame, name,(x1,y1-10),cv2.FONT_HERSHEY_DUPLEX,1,(0,0,0),2)
		cv2.rectangle(frame,(x1,y1),(x2,y2),(0,0,255),4)
		
		
	cv2.imshow("Frame",frame)
	key=cv2.waitKey(1)
	if key==27:
		lock(role_pini)
		break

cap.release()
cv2.destroyAllWindows()

data = pd.read_csv(day_str)

wb = openpyxl.Workbook()
sayfa = wb.active

a2 = len(data)           ### toplan satır sayısı
a3 = len(data.columns)   ### toplam sütun sayısı
#print('satır uzunluğu: ', a2)
#print('sütun sayısı: ', a3)


for x in range(a3):      ### sütun başlıklarını yazdırma döngüsü
    c = x + 1
    sayfa.cell(row = 1, column = c).value = data.columns[x]


for x in range(a2):    ### tüm satırlardaki verileri excele yazdırma döngüsü
    for y in range(a3):
        r = x + 2
        c = y + 1
        sayfa.cell(row = r, column = c).value = data.iat[x,y]


wb.save("yoklama-" + day + ".xlsx")
os.remove(day_str)
print('İşlem başarıyla tamamlandı. Excel dosyanız oluşturuldu')
